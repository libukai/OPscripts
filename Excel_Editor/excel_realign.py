#!/usr/bin/env python
# @Author: Libukai
# @Date:   2015-12-29 11:28:18
# @Last Modified by:   Libukai
# @Last Modified time: 2015-12-30 16:49:11


import os
import glob
import openpyxl

def get_data():
    '''
    从原始数据文件所在文件夹中，抽取数据文件对应日期，并以不可变的元组形式储存
    请将文件名统一调整为 xx_xx 的方式，例如 09_05.xlsx，以便生成正确的日期排序
    如果遇到单个的文件处理中报错，一般将文件重新另存为 .xlsx 格式的文件即可解决
    '''
    print('正在提取数据文件对应的日期……')
    file_list = glob.glob('*.xlsx')
    date_tuple = tuple([os.path.splitext(date)[0] for date in file_list])
    return date_tuple

def template_excel(keep_data_column_number):
    '''
    根据数据的实际情况，生成最终导出文件的保留栏位数据
    '''
    print('正在生成模版……')
    example_excel = openpyxl.load_workbook(glob.glob('*.xlsx')[0])
    example_sheet = example_excel.active
    template_excel = openpyxl.Workbook()
    template_sheet = template_excel.active
    keep_column_label = example_sheet.cell(row = 1, column = keep_data_column_number).column
    for example_row in example_sheet['A1':'%s%d' %(keep_column_label, len(example_sheet.rows))]:
        for cell in example_row:
            template_sheet['%s%d' %(cell.column, cell.row)] = cell.value
    return template_excel

def final_excel(template_excel, keep_data_column_number, input_data_column_number):
    '''
    循环遍历所有的数据文件，将需要的数据导入到模版文件中，并生成最终的Excel
    '''
    result_sheet = template_excel.active
    #初始化导出文件中的写入数据栏位
    count_column_number = 0    
    for data_date in get_data():
        print('正在处理日期为', data_date, '的数据……')
        input_excel = openpyxl.load_workbook('%s.xlsx' %data_date)
        input_sheet = input_excel.active
        count_column_number += 1
        #确定导出文件中的写入数据栏位的字母编号
        data_column = result_sheet.cell(row = 1, column = count_column_number + keep_data_column_number).column
        #初始化导出文件中的写入数据行数
        data_row = 0
        for input_row in input_sheet:
            data_row += 1
            #获取每一次迭代时，需要获取的单元格的值
            input_data = input_row[input_data_column_number - 1].value
            if data_row == 1:
                result_sheet['%s%d' %(data_column, 1)] = data_date
            else:
                if input_data == '\\N' or input_data == '':
                    result_sheet['%s%d' %(data_column, data_row)] = 0
                else:
                    result_sheet['%s%d' %(data_column, data_row)] = input_data
    print('正在生成最终的数据文件……')
    template_excel.save('output.xlsx')
        

if __name__ == '__main__':
    #定义导出文件中保留的通用数据列的截止栏位的数字编号
    keep_data_column_number = 3
    #定义数据文件中采用的数据列栏位的数字编号
    input_data_column_number = 4 
    final_excel(template_excel(keep_data_column_number), keep_data_column_number, input_data_column_number)
